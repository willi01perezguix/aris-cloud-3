from __future__ import annotations

import base64
import csv
import io
import json
import logging
import re
import time
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from typing import Any
from xml.etree import ElementTree

import httpx
from openpyxl import load_workbook

from app.aris3.core.config import settings
from app.aris3.core.error_catalog import AppError, ErrorCatalog


SUPPORTED_CONTENT_TYPES = {
    "image/jpeg",
    "image/png",
    "image/webp",
    "application/pdf",
    "text/csv",
    "text/tab-separated-values",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "text/plain",
}

EPC_KEYS = {"epc", "rfid", "tag"}
SALE_KEYS = {"venta", "sale", "sale_price", "precio venta"}
LOGISTICS_HINTS = {"en transito", "en camino", "entregado", "recibido", "enviado", "shipping", "delivery", "express"}
NON_SELLABLE_PHRASES = {
    "no vender",
    "not for sale",
    "muestra",
    "sample",
    "exhibición",
    "display only",
    "dañado",
    "damaged",
    "defectuoso",
    "defective",
    "cancelado",
    "canceled",
    "reembolsado",
    "refunded",
    "faltante",
    "missing",
    "devuelto",
    "returned",
    "uso interno",
    "internal use",
}
DEFAULT_OPENAI_MODEL = "gpt-4.1-mini"
_MODEL_NAME_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._:-]{1,127}$")
OPENAI_TOTAL_TIMEOUT_SECONDS = 18.0
OPENAI_CONNECT_TIMEOUT_SECONDS = 3.0
logger = logging.getLogger(__name__)

INVENTORY_PRELOAD_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "document_summary": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "document_type": {"type": ["string", "null"]},
                "document_number": {"type": ["string", "null"]},
                "document_date": {"type": ["string", "null"]},
                "detected_currency": {"type": ["string", "null"]},
                "overall_confidence": {"type": ["number", "null"]},
            },
            "required": ["document_type", "document_number", "document_date", "detected_currency", "overall_confidence"],
        },
        "pricing_context": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "source_currency": {"type": ["string", "null"]},
                "exchange_rate_to_gtq": {"type": ["string", "null"]},
                "pricing_mode": {"type": ["string", "null"]},
            },
            "required": ["source_currency", "exchange_rate_to_gtq", "pricing_mode"],
        },
        "lines": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "sku": {"type": ["string", "null"]},
                    "suggested_sku": {"type": ["string", "null"]},
                    "description": {"type": ["string", "null"]},
                    "variant_1": {"type": ["string", "null"]},
                    "variant_2": {"type": ["string", "null"]},
                    "brand": {"type": ["string", "null"]},
                    "category": {"type": ["string", "null"]},
                    "style": {"type": ["string", "null"]},
                    "color": {"type": ["string", "null"]},
                    "size": {"type": ["string", "null"]},
                    "pool": {"type": ["string", "null"]},
                    "location_code": {"type": ["string", "null"]},
                    "logistics_status": {"type": ["string", "null"]},
                    "sellable": {"type": ["boolean", "null"]},
                    "quantity": {"type": ["integer", "null"]},
                    "original_cost": {"type": ["string", "null"]},
                    "source_currency": {"type": ["string", "null"]},
                    "needs_review": {"type": ["boolean", "null"]},
                    "confidence": {"type": ["number", "null"]},
                    "notes": {"type": ["string", "null"]},
                    "source_order_number": {"type": ["string", "null"]},
                    "source_order_date": {"type": ["string", "null"]},
                    "source_supplier": {"type": ["string", "null"]},
                    "reference_price_original": {"type": ["string", "null"]},
                    "reference_price_gtq": {"type": ["string", "null"]},
                    "source_file_name": {"type": ["string", "null"]},
                    "source_row_number": {"type": ["integer", "null"]},
                },
                "required": [
                    "sku",
                    "suggested_sku",
                    "description",
                    "variant_1",
                    "variant_2",
                    "brand",
                    "category",
                    "style",
                    "color",
                    "size",
                    "pool",
                    "location_code",
                    "logistics_status",
                    "sellable",
                    "quantity",
                    "original_cost",
                    "source_currency",
                    "needs_review",
                    "confidence",
                    "notes",
                    "source_order_number",
                    "source_order_date",
                    "source_supplier",
                    "reference_price_original",
                    "reference_price_gtq",
                    "source_file_name",
                    "source_row_number",
                ],
            },
        },
        "warnings": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "severity": {"type": ["string", "null"]},
                    "message": {"type": ["string", "null"]},
                    "line_ref": {"type": ["string", "null"]},
                },
                "required": ["severity", "message", "line_ref"],
            },
        },
    },
    "required": ["document_summary", "pricing_context", "lines", "warnings"],
}


@dataclass
class UploadedSource:
    filename: str
    content_type: str
    content: bytes


class OpenAIInventoryClient:
    def __init__(self) -> None:
        self._api_key = settings.OPENAI_API_KEY
        configured_model = (settings.OPENAI_INVENTORY_MODEL or "").strip()
        if configured_model and _MODEL_NAME_PATTERN.match(configured_model):
            self._model = configured_model
            self._invalid_configured_model: str | None = None
        elif configured_model:
            self._model = configured_model
            self._invalid_configured_model = configured_model
        else:
            self._model = DEFAULT_OPENAI_MODEL
            self._invalid_configured_model = None
        logger.info(
            "stock.ai_preload.config api_key_present=%s model_configured=%s model_used=%s",
            bool(self._api_key),
            bool(configured_model),
            self._model,
        )

    def extract(
        self,
        *,
        prompt: str,
        attachments: list[UploadedSource],
        trace_id: str | None,
        tenant_id: str,
        store_id: str,
        document_type: str | None,
        timeout_seconds: float | None = None,
    ) -> dict[str, Any]:
        if not self._api_key:
            raise AppError(
                ErrorCatalog.VALIDATION_ERROR,
                details={"message": "OPENAI_API_KEY is not configured", "retryable": False, "model": self._model},
            )
        if self._invalid_configured_model:
            raise AppError(
                ErrorCatalog.AI_INVALID_MODEL,
                details={"message": "Configured OpenAI model is invalid", "retryable": False, "model": self._invalid_configured_model},
            )
        content: list[dict[str, Any]] = [{"type": "input_text", "text": prompt}]
        for item in attachments:
            encoded = base64.b64encode(item.content).decode("ascii")
            if item.content_type.startswith("image/"):
                content.append({"type": "input_image", "image_url": f"data:{item.content_type};base64,{encoded}"})
            else:
                content.append({"type": "input_file", "filename": item.filename, "file_data": f"data:{item.content_type};base64,{encoded}"})

        payload = {
            "model": self._model,
            "input": [{"role": "system", "content": [{"type": "input_text", "text": "Extract inventory lines."}]}, {"role": "user", "content": content}],
            "text": {
                "format": {
                    "type": "json_schema",
                    "name": "inventory_preload_extraction",
                    "strict": True,
                    "schema": INVENTORY_PRELOAD_SCHEMA,
                }
            },
        }

        total_timeout = timeout_seconds if timeout_seconds is not None else OPENAI_TOTAL_TIMEOUT_SECONDS
        timeout = httpx.Timeout(total_timeout, connect=OPENAI_CONNECT_TIMEOUT_SECONDS)
        start = time.perf_counter()
        logger.info(
            "stock.ai_preload.openai_call_started trace_id=%s tenant_id=%s store_id=%s document_type=%s text_only=%s files_count=%s model_used=%s",
            trace_id,
            tenant_id,
            store_id,
            document_type,
            len(attachments) == 0,
            len(attachments),
            self._model,
        )
        try:
            response = httpx.post(
                "https://api.openai.com/v1/responses",
                headers={"Authorization": f"Bearer {self._api_key}"},
                json=payload,
                timeout=timeout,
            )
            response.raise_for_status()
            data = response.json()
        except httpx.TimeoutException as exc:
            elapsed_ms = int((time.perf_counter() - start) * 1000)
            logger.warning(
                "stock.ai_preload.openai_call_finished trace_id=%s tenant_id=%s store_id=%s success=false timeout=true elapsed_ms=%s error_class=%s model_used=%s",
                trace_id,
                tenant_id,
                store_id,
                elapsed_ms,
                exc.__class__.__name__,
                self._model,
            )
            raise AppError(
                ErrorCatalog.AI_SERVICE_TIMEOUT,
                details={"message": "AI inventory analysis timed out", "retryable": True, "model": self._model},
            ) from exc
        except httpx.HTTPStatusError as exc:
            elapsed_ms = int((time.perf_counter() - start) * 1000)
            status_code = exc.response.status_code if exc.response is not None else None
            provider_error = _extract_openai_error(exc.response)
            error_lower = (provider_error.get("message") or "").lower()
            if status_code in (401, 403):
                error = ErrorCatalog.AI_AUTH_FAILED
                details = {"message": "OpenAI API key is invalid or unauthorized", "retryable": False, "model": self._model}
            elif status_code == 429:
                error = ErrorCatalog.AI_RATE_LIMITED
                details = {"message": "OpenAI rate limit exceeded", "retryable": True, "model": self._model}
            elif status_code == 400 and ("model" in error_lower and ("not found" in error_lower or "does not exist" in error_lower)):
                error = ErrorCatalog.AI_INVALID_MODEL
                details = {
                    "message": "Configured OpenAI model is invalid",
                    "retryable": False,
                    "model": self._model,
                    "provider_status": status_code,
                    "provider_error_type": provider_error.get("type"),
                    "provider_error_code": provider_error.get("code"),
                    "provider_error_param": provider_error.get("param"),
                }
            elif status_code == 400:
                error = ErrorCatalog.AI_REQUEST_INVALID
                details = {
                    "message": "AI inventory request is invalid",
                    "retryable": False,
                    "model": self._model,
                    "provider_status": status_code,
                    "provider_error_type": provider_error.get("type"),
                    "provider_error_code": provider_error.get("code"),
                    "provider_error_param": provider_error.get("param"),
                }
            elif status_code and status_code >= 500:
                error = ErrorCatalog.AI_SERVICE_UNAVAILABLE
                details = {
                    "message": "OpenAI service unavailable",
                    "retryable": True,
                    "model": self._model,
                    "provider_status": status_code,
                    "provider_error_type": provider_error.get("type"),
                    "provider_error_code": provider_error.get("code"),
                    "provider_error_param": provider_error.get("param"),
                }
            else:
                error = ErrorCatalog.AI_SERVICE_UNAVAILABLE
                details = {
                    "message": "AI analysis failed",
                    "retryable": status_code != 400,
                    "model": self._model,
                    "provider_status": status_code,
                    "provider_error_type": provider_error.get("type"),
                    "provider_error_code": provider_error.get("code"),
                    "provider_error_param": provider_error.get("param"),
                }
            logger.warning(
                "stock.ai_preload.openai_call_finished trace_id=%s tenant_id=%s store_id=%s success=false timeout=false elapsed_ms=%s error_class=%s status_code=%s model_used=%s provider_error_type=%s provider_error_code=%s provider_error_param=%s provider_error_message=%s",
                trace_id,
                tenant_id,
                store_id,
                elapsed_ms,
                exc.__class__.__name__,
                status_code,
                self._model,
                provider_error.get("type"),
                provider_error.get("code"),
                provider_error.get("param"),
                provider_error.get("message"),
            )
            raise AppError(error, details=details) from exc
        except httpx.HTTPError as exc:
            elapsed_ms = int((time.perf_counter() - start) * 1000)
            logger.warning(
                "stock.ai_preload.openai_call_finished trace_id=%s tenant_id=%s store_id=%s success=false timeout=false elapsed_ms=%s error_class=%s model_used=%s",
                trace_id,
                tenant_id,
                store_id,
                elapsed_ms,
                exc.__class__.__name__,
                self._model,
            )
            raise AppError(
                ErrorCatalog.AI_SERVICE_UNAVAILABLE,
                details={"message": "AI analysis failed", "retryable": True, "model": self._model},
            ) from exc

        elapsed_ms = int((time.perf_counter() - start) * 1000)
        logger.info(
            "stock.ai_preload.openai_call_finished trace_id=%s tenant_id=%s store_id=%s success=true timeout=false elapsed_ms=%s model_used=%s",
            trace_id,
            tenant_id,
            store_id,
            elapsed_ms,
            self._model,
        )

        text_output = data.get("output_text")
        if text_output:
            try:
                return json.loads(text_output)
            except json.JSONDecodeError as exc:
                raise AppError(
                    ErrorCatalog.AI_RESPONSE_INVALID,
                    details={"message": "OpenAI returned malformed JSON output", "retryable": False, "model": self._model},
                ) from exc
        output = data.get("output", [])
        for item in output:
            for content_item in item.get("content", []):
                if content_item.get("type") == "output_text":
                    try:
                        return json.loads(content_item.get("text", "{}"))
                    except json.JSONDecodeError as exc:
                        raise AppError(
                            ErrorCatalog.AI_RESPONSE_INVALID,
                            details={"message": "OpenAI returned malformed JSON output", "retryable": False, "model": self._model},
                        ) from exc
        raise AppError(
            ErrorCatalog.AI_RESPONSE_INVALID,
            details={"message": "AI analysis returned empty output", "retryable": False, "model": self._model},
        )


def _extract_openai_error(response: httpx.Response | None) -> dict[str, Any]:
    if response is None:
        return {"message": None, "type": None, "param": None, "code": None}
    payload: dict[str, Any] = {}
    try:
        payload = response.json() if response.content else {}
    except (ValueError, TypeError):
        payload = {}
    error_payload = payload.get("error") if isinstance(payload, dict) else {}
    if not isinstance(error_payload, dict):
        error_payload = {}
    return {
        "message": error_payload.get("message"),
        "type": error_payload.get("type"),
        "param": error_payload.get("param"),
        "code": error_payload.get("code"),
    }


class StockAiPreloadService:
    def __init__(self, openai_client: OpenAIInventoryClient | None = None) -> None:
        self.openai_client = openai_client or OpenAIInventoryClient()

    def validate_files(self, files: list[UploadedSource]) -> None:
        if len(files) > settings.AI_PRELOAD_MAX_FILES:
            raise AppError(ErrorCatalog.VALIDATION_ERROR, details={"message": "too many files"})
        total_bytes = 0
        for file in files:
            if file.content_type not in SUPPORTED_CONTENT_TYPES:
                raise AppError(ErrorCatalog.VALIDATION_ERROR, details={"message": f"unsupported file type: {file.content_type}"})
            if len(file.content) > settings.AI_PRELOAD_MAX_FILE_BYTES:
                raise AppError(ErrorCatalog.VALIDATION_ERROR, details={"message": f"file too large: {file.filename}"})
            total_bytes += len(file.content)
        if total_bytes > settings.AI_PRELOAD_MAX_TOTAL_BYTES:
            raise AppError(ErrorCatalog.VALIDATION_ERROR, details={"message": "total upload size exceeded"})

    def extract_spreadsheet_rows(self, file: UploadedSource) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
        warnings: list[dict[str, str]] = []
        rows: list[dict[str, str]] = []
        if file.content_type in {"text/csv", "text/tab-separated-values"}:
            delimiter = "\t" if file.content_type == "text/tab-separated-values" else ","
            reader = csv.DictReader(io.StringIO(file.content.decode("utf-8", errors="ignore")), delimiter=delimiter)
            for idx, row in enumerate(reader, start=2):
                normalized = {str(k or "").strip().lower(): str(v or "").strip() for k, v in row.items()}
                normalized["_row_number"] = str(idx)
                rows.append(normalized)
        else:
            wb = load_workbook(io.BytesIO(file.content), data_only=True)
            for sheet in wb.worksheets:
                values = list(sheet.iter_rows(values_only=True))
                if not values:
                    continue
                headers = [str(v or "").strip().lower() for v in values[0]]
                for idx, value_row in enumerate(values[1:], start=2):
                    normalized = {headers[i]: str(value_row[i] or "").strip() for i in range(len(headers)) if headers[i]}
                    normalized["_row_number"] = str(idx)
                    normalized["_sheet_name"] = sheet.title
                    rows.append(normalized)
        if not rows:
            return rows, warnings
        keys = {k for row in rows for k in row.keys()}
        if any(any(token in key for token in EPC_KEYS) for key in keys):
            warnings.append({"severity": "info", "message": "EPC values were detected but ignored. EPC must be assigned later."})
        if any(any(token in key for token in SALE_KEYS) for key in keys):
            warnings.append({"severity": "info", "message": "Sale price values were detected but not imported. Final sale price must be set later."})
        return rows, warnings

    def map_deterministic_row(self, row: dict[str, str], *, source_file_name: str | None = None) -> tuple[dict[str, Any], list[str]]:
        warnings: list[str] = []
        get = lambda *keys: next((row.get(k, "").strip() for k in keys if row.get(k, "").strip()), "")
        sku = get("sku", "skc", "item code", "codigo", "código")
        description = get("descripción", "descripcion", "description", "product title")
        variant_1 = get("color", "variante 1", "variant_1")
        variant_2 = get("talla", "size", "variante 2", "variant_2")
        quantity_raw = get("cantidad", "qty", "quantity")
        quantity = int(quantity_raw) if quantity_raw.isdigit() else 1
        line: dict[str, Any] = {
            "sku": sku or None,
            "description": description or "",
            "variant_1": variant_1 or None,
            "variant_2": variant_2 or None,
            "color": variant_1 or None,
            "size": variant_2 or None,
            "brand": get("marca", "brand") or None,
            "category": get("categoría", "categoria", "category") or None,
            "style": get("estilo", "style") or None,
            "source_order_number": get("número de pedido", "numero de pedido", "order number") or None,
            "source_order_date": get("fecha pedido", "fecha", "order date") or None,
            "source_supplier": get("supplier", "proveedor", "supplier name") or None,
            "pool": "BODEGA",
            "location_code": "RECEPCION",
            "logistics_status": None,
            "sellable": True,
            "quantity": max(1, quantity),
            "needs_review": False,
            "source_file_name": source_file_name,
            "source_row_number": int((row.get("_row_number") or "0") or "0") or None,
        }
        precio_costo_gtq = get("precio costo (q)", "costo q", "costo gtq", "quetzales")
        precio_usd = get("precio (usd)", "precio usd", "usd", "unit price(usd)")
        precio_final = get("precio final (q)", "precio final")
        precio_sugerido = get("precio venta sugerido (q)", "precio sugerido", "suggested price")
        location_raw = get("ubicación", "ubicacion", "location", "estado envío", "estado envio")
        if location_raw:
            lowered = location_raw.lower()
            if any(hint in lowered for hint in LOGISTICS_HINTS):
                line["logistics_status"] = location_raw.upper().replace(" ", "_")
                line["notes"] = f"Logistics source value preserved: {location_raw}"
                warnings.append("logistics location not mapped to ARIS location")
            elif re.match(r"^[A-Za-z0-9_-]{2,24}$", location_raw):
                line["location_code"] = location_raw.upper()
            else:
                line["notes"] = f"Location source value preserved: {location_raw}"
                warnings.append("logistics location not mapped to ARIS location")

        if precio_costo_gtq:
            line["cost_gtq"] = str(self.parse_decimal(precio_costo_gtq))
            line["source_currency"] = "GTQ"
            line["exchange_rate_to_gtq"] = "1.00"
        if precio_usd:
            line["original_cost"] = str(self.parse_decimal(precio_usd))
            line["source_currency"] = "USD"
        symbol_price = get("precio", "price")
        if symbol_price.startswith("$") and not line.get("source_currency"):
            line["source_currency"] = "unknown"
            line["needs_review"] = True
            warnings.append("Confirmar moneda antes de calcular Costo(Q).")
        if precio_final:
            line["reference_price_gtq"] = str(self.parse_decimal(precio_final))
            warnings.append("final sale price detected but not imported")
        if precio_sugerido:
            line["suggested_price_gtq"] = str(self.parse_decimal(precio_sugerido))

        if line.get("original_cost") and line.get("cost_gtq"):
            try:
                implied = (Decimal(line["cost_gtq"]) / Decimal(line["original_cost"])).quantize(Decimal("0.01"))
                line["exchange_rate_to_gtq"] = format(implied, "f")
            except (InvalidOperation, ZeroDivisionError):
                line["needs_review"] = True

        self.apply_sellable_rules(line)
        if not line.get("sku"):
            line["needs_review"] = True
            warnings.append("missing SKU")
        if not line.get("quantity"):
            line["needs_review"] = True
            warnings.append("missing quantity")
        if not (line.get("original_cost") or line.get("cost_gtq")):
            line["needs_review"] = True
            warnings.append("missing cost")
        return line, warnings

    def apply_sellable_rules(self, line: dict[str, Any]) -> None:
        notes = (line.get("notes") or "").lower()
        description = (line.get("description") or "").lower()
        joined = f"{notes} {description}"
        hit = next((phrase for phrase in NON_SELLABLE_PHRASES if phrase in joined), None)
        if hit:
            line["sellable"] = False
            line["needs_review"] = True
            line["notes"] = (line.get("notes") or "") + f" Explicit non-sellable phrase: {hit}"
        else:
            line["sellable"] = True

    def extract_docx_text(self, content: bytes) -> str:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            xml_data = archive.read("word/document.xml")
        root = ElementTree.fromstring(xml_data)
        return "\n".join(node.text for node in root.iter() if node.text)

    def parse_decimal(self, value: str | None) -> Decimal | None:
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError) as exc:
            raise AppError(ErrorCatalog.VALIDATION_ERROR, details={"message": f"invalid decimal: {value}"}) from exc

    def round_up(self, value: Decimal, step: Decimal) -> Decimal:
        return (value / step).to_integral_value(rounding=ROUND_CEILING) * step

    def apply_pricing(
        self,
        *,
        line: dict[str, Any],
        source_currency: str,
        exchange_rate_to_gtq: Decimal | None,
        pricing_mode: str,
        markup_percent: Decimal | None,
        margin_percent: Decimal | None,
        multiplier: Decimal | None,
        rounding_step: Decimal,
    ) -> dict[str, Any]:
        needs_review = bool(line.get("needs_review", False))
        original_cost = self.parse_decimal(line.get("original_cost"))
        existing_cost_gtq = self.parse_decimal(line.get("cost_gtq"))
        reference_price_original = self.parse_decimal(line.get("reference_price_original"))
        reference_price_gtq = self.parse_decimal(line.get("reference_price_gtq"))
        line_currency = (line.get("source_currency") or source_currency or "GTQ").upper()
        if original_cost is not None:
            if line_currency == "GTQ":
                rate = Decimal("1.00")
            else:
                rate = exchange_rate_to_gtq
                if rate is None:
                    needs_review = True
            if existing_cost_gtq is not None:
                cost_gtq = existing_cost_gtq
                if original_cost > 0:
                    implied_rate = (cost_gtq / original_cost).quantize(Decimal("0.01"))
                    line["exchange_rate_to_gtq"] = format(implied_rate, "f")
            elif rate is not None:
                cost_gtq = (original_cost * rate).quantize(Decimal("0.01"))
                line["cost_gtq"] = format(cost_gtq, "f")
                line["exchange_rate_to_gtq"] = format(rate, "f")
            else:
                cost_gtq = None
            if cost_gtq is not None:
                line["cost_gtq"] = format(cost_gtq, "f")
                if pricing_mode == "markup_percent" and markup_percent is not None:
                    base = cost_gtq * (Decimal("1") + (markup_percent / Decimal("100")))
                    line["suggested_price_gtq"] = format(self.round_up(base, rounding_step).quantize(Decimal("0.01")), "f")
                elif pricing_mode == "margin_percent" and margin_percent is not None:
                    base = cost_gtq / (Decimal("1") - (margin_percent / Decimal("100")))
                    line["suggested_price_gtq"] = format(self.round_up(base, rounding_step).quantize(Decimal("0.01")), "f")
                elif pricing_mode == "multiplier" and multiplier is not None:
                    base = cost_gtq * multiplier
                    line["suggested_price_gtq"] = format(self.round_up(base, rounding_step).quantize(Decimal("0.01")), "f")
                elif pricing_mode == "manual":
                    line["suggested_price_gtq"] = line.get("suggested_price_gtq")
                    needs_review = True
        if reference_price_original is not None:
            if line_currency == "GTQ":
                line["reference_price_gtq"] = format(reference_price_original, "f")
            elif exchange_rate_to_gtq is not None:
                computed_reference_gtq = (reference_price_original * exchange_rate_to_gtq).quantize(Decimal("0.01"))
                line["reference_price_gtq"] = format(computed_reference_gtq, "f")
        if reference_price_gtq is not None:
            line["reference_price_gtq"] = format(reference_price_gtq, "f")
        line["source_currency"] = line_currency
        line["needs_review"] = needs_review
        return line

    def apply_operational_defaults(self, line: dict[str, Any]) -> dict[str, Any]:
        if not line.get("pool"):
            line["pool"] = "BODEGA"
        if not line.get("location_code"):
            line["location_code"] = "RECEPCION"
        return line

    def apply_shein_reference_price_hint(self, *, line: dict[str, Any], free_text: str | None) -> bool:
        if line.get("reference_price_original") is not None or not free_text:
            return False
        amounts = re.findall(r"\$(\d+(?:\.\d{1,2})?)", free_text)
        if len(amounts) < 2:
            return False
        first_amount = self.parse_decimal(amounts[0])
        second_amount = self.parse_decimal(amounts[1])
        if first_amount is None or second_amount is None:
            return False
        original_cost = self.parse_decimal(line.get("original_cost"))
        if original_cost is None or original_cost != first_amount:
            return False
        line["reference_price_original"] = format(second_amount, "f")
        return True

    def is_large_text_input(self, free_text: str | None) -> bool:
        if not free_text:
            return False
        text = free_text.strip()
        if not text:
            return False
        if len(text) > settings.AI_PRELOAD_SYNC_TEXT_CHAR_LIMIT:
            return True
        if text.count("SKU:") >= 5:
            return True
        lowered = text.lower()
        return "núm. de pedido" in lowered and "productos" in lowered

    def parse_shein_order_text(self, *, free_text: str | None, source_currency: str) -> dict[str, Any] | None:
        if not free_text:
            return None
        lowered = free_text.lower()
        if "sku:" not in lowered or ("shein" not in lowered and "núm. de pedido" not in lowered):
            return None

        order_number = self._extract_order_number(free_text)
        order_date = self._extract_order_date(free_text)
        sku_matches = list(re.finditer(r"(\d+)\s+SKU:\s*([A-Za-z0-9_-]+)", free_text, flags=re.IGNORECASE))
        if not sku_matches:
            return None

        lines: list[dict[str, Any]] = []
        for idx, match in enumerate(sku_matches):
            quantity = int(match.group(1))
            sku = match.group(2)
            prev_start = sku_matches[idx - 1].end() if idx > 0 else 0
            block_before = free_text[max(prev_start, match.start() - 420):match.start()]
            block_after = free_text[match.end(): sku_matches[idx + 1].start() if idx + 1 < len(sku_matches) else min(len(free_text), match.end() + 240)]

            description, variant_1, variant_2 = self._extract_shein_description_and_variants(block_before)
            amounts = re.findall(r"\$(\d+(?:\.\d{1,2})?)", block_after)
            if len(amounts) < 2:
                amounts = re.findall(r"\$(\d+(?:\.\d{1,2})?)", block_before + "\n" + block_after)
            first_price = self.parse_decimal(amounts[0]) if len(amounts) >= 1 else None
            second_price = self.parse_decimal(amounts[1]) if len(amounts) >= 2 else None
            status_match = re.search(r"\b(Enviado|Entregado|Recibido|En tránsito|En transito)\b", block_after, flags=re.IGNORECASE)
            logistics_status = status_match.group(1) if status_match else None

            line: dict[str, Any] = {
                "sku": sku,
                "description": description or f"Producto SHEIN {sku}",
                "variant_1": variant_1,
                "variant_2": variant_2,
                "color": variant_1,
                "size": variant_2,
                "brand": "SHEIN" if "shein" in (description or "").lower() else None,
                "pool": "BODEGA",
                "location_code": "RECEPCION",
                "logistics_status": logistics_status or "Enviado",
                "sellable": True,
                "quantity": max(1, quantity),
                "source_order_number": order_number,
                "source_order_date": order_date,
                "source_currency": source_currency,
                "original_cost": format(first_price, "f") if first_price is not None else None,
                "reference_price_original": format(second_price, "f") if second_price is not None else None,
                "needs_review": first_price is None,
                "confidence": 0.93,
                "notes": None,
            }
            lines.append(line)

        high_confidence_lines = [line for line in lines if line.get("sku") and line.get("original_cost") and line.get("reference_price_original")]
        if not high_confidence_lines:
            return None
        return {
            "document_summary": {
                "document_type": "other",
                "document_number": order_number,
                "document_date": order_date,
                "detected_currency": source_currency,
                "overall_confidence": 0.92,
            },
            "lines": lines,
            "warnings": [],
        }

    def _extract_order_number(self, text: str) -> str | None:
        inline = re.search(r"N[úu]m\.\s*de\s*pedido\s*[:\s]+([A-Z0-9]{8,})", text, flags=re.IGNORECASE)
        if inline:
            return inline.group(1).strip()
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        for idx, line in enumerate(lines[:-1]):
            if re.search(r"N[úu]m\.\s*de\s*pedido", line, flags=re.IGNORECASE):
                candidate = lines[idx + 1].strip()
                if re.fullmatch(r"[A-Z0-9]{8,}", candidate):
                    return candidate
        return None

    def _extract_order_date(self, text: str) -> str | None:
        inline = re.search(r"Fecha\s*[:\s]+([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{4}|[0-9]{4}-[0-9]{2}-[0-9]{2})", text, flags=re.IGNORECASE)
        if not inline:
            return None
        raw = inline.group(1).strip()
        iso = re.fullmatch(r"([0-9]{4}-[0-9]{2}-[0-9]{2})", raw)
        if iso:
            return iso.group(1)
        parts = raw.split()
        if len(parts) != 3:
            return raw
        month_map = {
            "jan": "01", "ene": "01", "feb": "02", "mar": "03", "apr": "04", "abr": "04", "may": "05", "jun": "06",
            "jul": "07", "aug": "08", "ago": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12", "dic": "12",
        }
        day, month, year = parts
        month_num = month_map.get(month.lower()[:3])
        if not month_num:
            return raw
        return f"{year}-{month_num}-{int(day):02d}"

    def _extract_shein_description_and_variants(self, block_before: str) -> tuple[str | None, str | None, str | None]:
        lines = [line.strip() for line in block_before.splitlines() if line.strip()]
        variant_1 = None
        variant_2 = None
        description = None
        for line in reversed(lines):
            if "/" in line and len(line) <= 80:
                parts = [part.strip() for part in line.split("/", 1)]
                if len(parts) == 2:
                    variant_1, variant_2 = parts[0] or None, parts[1] or None
                    break
        for line in reversed(lines):
            lowered = line.lower()
            if "productos" in lowered or "cantidad" in lowered or "sku" in lowered:
                continue
            if "/" in line and line == f"{variant_1 or ''} / {variant_2 or ''}".strip():
                continue
            if len(line) > 8:
                description = line
                break
        return description, variant_1, variant_2
